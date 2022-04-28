from openpyxl import load_workbook, workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import inflection
import string


def main():
    newWorkbook = workbook.Workbook()
    # style for tables
    style = TableStyleInfo(
        name="TableStyleMedium5",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    del newWorkbook["Sheet"]
    workBook = load_workbook("./files/input/tus-2022-1-converted.xlsx")
    # translation rule to remove turkish characters from table and sheet names
    Tr2Eng = str.maketrans("çğıöşü", "cgiosu")
    # translation rule to remove punctuation from table and sheet names
    NoPunc = str.maketrans("", "", string.punctuation)
    # dictionary to store departments uniquely
    departments = {}
    # firstpage has an additional row for the header
    firstPage = True
    # loop through all sheets of the converted result file
    for sheet in workBook.worksheets:
        maxColumn = sheet.max_column
        maxRow = sheet.max_row
        for i in range(3 if firstPage else 2, maxRow + 1):
            # specialization area names are in the column 2 (B) of the sheets
            cellData = sheet.cell(row=i, column=2).value
            if cellData:
                deptName = inflection.camelize(
                    cellData.split("/")[-1].lower().translate(NoPunc).translate(Tr2Eng).replace(" ", "_")
                )[:30]
                # add area name to dictionary and create a new sheet for it
                if deptName not in departments:
                    newWorkbook.create_sheet(deptName)
                    departments[deptName] = {
                        "name": cellData.split("/")[-1],
                        "rowIndex": 2,
                        "generalTotalQuota": 0,
                        "generalFilledQuota": 0,
                        "generalEmptyQuota": 0,
                        "foreignTotalQuota": 0,
                        "foreignFilledQuota": 0,
                        "foreignEmptyQuota": 0,
                    }
                # copy existing data of the area to the new sheet
                for k in range(1, maxColumn + 1):
                    newWorkbook[deptName].cell(row=departments[deptName]["rowIndex"], column=k).value = sheet.cell(
                        row=i, column=k
                    ).value
                # update quota values
                if "yabancı" in sheet.cell(row=i, column=3).value.lower():
                    departments[deptName]["foreignTotalQuota"] += int(sheet.cell(row=i, column=4).value)
                    departments[deptName]["foreignFilledQuota"] += int(sheet.cell(row=i, column=5).value)
                    departments[deptName]["foreignEmptyQuota"] += int(sheet.cell(row=i, column=6).value)
                else:
                    departments[deptName]["generalTotalQuota"] += int(sheet.cell(row=i, column=4).value)
                    departments[deptName]["generalFilledQuota"] += int(sheet.cell(row=i, column=5).value)
                    departments[deptName]["generalEmptyQuota"] += int(sheet.cell(row=i, column=6).value)
                departments[deptName]["rowIndex"] += 1
        firstPage = False

    # create a new sheet for the total quota information of all departments
    newWorkbook.create_sheet("Kontenjanlar")
    newWorkbook["Kontenjanlar"].cell(row=1, column=1).value = "Bölüm"
    newWorkbook["Kontenjanlar"].cell(row=1, column=2).value = "Toplam Kontenjan (Genel)"
    newWorkbook["Kontenjanlar"].cell(row=1, column=3).value = "Yerleşen Sayısı (Genel)"
    newWorkbook["Kontenjanlar"].cell(row=1, column=4).value = "Boş Kalan Kontenjan (Genel)"
    newWorkbook["Kontenjanlar"].cell(row=1, column=5).value = "Toplam Kontenjan (Yabancı Uyruk)"
    newWorkbook["Kontenjanlar"].cell(row=1, column=6).value = "Yerleşen Sayısı (Yabancı Uyruk)"
    newWorkbook["Kontenjanlar"].cell(row=1, column=7).value = "Boş Kalan Kontenjan (Yabancı Uyruk)"
    rowIndex = 1
    for deptKey in departments.keys():
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=1).value = departments[deptKey]["name"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=2).value = departments[deptKey]["generalTotalQuota"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=3).value = departments[deptKey]["generalFilledQuota"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=4).value = departments[deptKey]["generalEmptyQuota"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=5).value = departments[deptKey]["foreignTotalQuota"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=6).value = departments[deptKey]["foreignFilledQuota"]
        newWorkbook["Kontenjanlar"].cell(row=rowIndex + 1, column=7).value = departments[deptKey]["foreignEmptyQuota"]
        rowIndex += 1

    for worksheet in newWorkbook.worksheets:
        if worksheet.title != "Kontenjanlar":
            # add column titles to all worksheets
            for i in range(1, worksheet.max_column + 1):
                worksheet.cell(row=1, column=i).value = workBook.worksheets[0].cell(row=2, column=i).value
            # add total quota and filled quota to the last row of each worksheet
            lastRow = worksheet.max_row + 1
            worksheet.cell(row=lastRow, column=3).value = "Genel Kontenjan"
            worksheet.cell(row=lastRow, column=4).value = departments[worksheet.title]["generalTotalQuota"]
            worksheet.cell(row=lastRow, column=5).value = departments[worksheet.title]["generalFilledQuota"]
            worksheet.cell(row=lastRow, column=6).value = departments[worksheet.title]["generalEmptyQuota"]
            worksheet.cell(row=lastRow + 1, column=3).value = "Yabancı Kontenjan"
            worksheet.cell(row=lastRow + 1, column=4).value = departments[worksheet.title]["foreignTotalQuota"]
            worksheet.cell(row=lastRow + 1, column=5).value = departments[worksheet.title]["foreignFilledQuota"]
            worksheet.cell(row=lastRow + 1, column=6).value = departments[worksheet.title]["foreignEmptyQuota"]
        # add table formatting and table styles to worksheets
        newTable = Table(
            displayName=worksheet.title,
            ref="A1:%s%s" % ("H" if worksheet.title != "Kontenjanlar" else "G", worksheet.max_row),
        )
        newTable.tableStyleInfo = style
        newWorkbook[worksheet.title].add_table(newTable)
        # adjust column widths in worksheets
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

    newWorkbook.save("./files/output/tus-yerlestirme-sonuclari-2022-1.xlsx")


main()
